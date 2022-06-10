#!/usr/bin/env python
# -*- coding: utf-8 -*-
# @Time    : 2022/6/6 20:10
# @File    : xunjian.py
# @Software: PyCharm
# @Description :

import logging
# import openpyxl
import os
import os.path
import pandas
from openpyxl.reader.excel import load_workbook
from netmiko import ConnectHandler
from netmiko.ssh_exception import (NetMikoTimeoutException, AuthenticationException, SSHException)
from multiprocessing.pool import ThreadPool
from datetime import datetime

class XunJian(object):
	def __init__(self):
		"""初始参数"""
		self.device_file = '设备信息表.xlsx'
		self.pool = ThreadPool(10)
		self.log ='XJJG'
		if not os.path.exists(self.log): os.mkdir(self.log)
		self.logtime = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
	def write_to_file(self, **kwargs):
		"""将结果写入文件"""
		if kwargs['state'] == 1:
			# 将正常的写入文件
			with open(kwargs['path'], 'a') as f:
				f.write(kwargs['result'])
		elif kwargs['state'] == 2:
			# 连接测试结果写入文件
			with open(os.path.join(self.log, f'connect_t_{self.logtime}.log'), 'a') as f:
				f.write(kwargs['result'])
				f.write('\n')
		else:
			# 将异常的写入文件
			with open(os.path.join(self.log, f'error_{self.logtime}.log'), 'a') as f:
				f.write(kwargs['error'])
				f.write('\n')
	def load_excel(self):
		"""
		加载excel文件
		:return:
		"""
		try:
			wb = load_workbook(self.device_file)
			return wb
		except FileNotFoundError:
			print('{}文件不存在'.format(self.device_file))
	def get_device_info(self):
		"""
		获取设备基本信息：
		:return:
		"""
		try:
			# 方法1：openpyxl
			wb = self.load_excel()
			ws1 = wb[wb.sheetnames[0]]
			# 通过参数min_row、max_col限制区域
			for row in ws1.iter_rows(min_row=2, max_col=9):
				if str(row[1].value).strip() == '#':
					# 跳过注释行
					continue
				info_dict = {'ip': row[2].value,
				             'protocol': row[3].value,
				             'port': row[4].value,
				             'username': row[5].value,
				             'password': row[6].value,
				             'secret': row[7].value,
				             'device_type': row[8].value,
				             'cmd_list': self.get_cmd_info(wb[row[8].value.strip().lower()]),
				             }
				yield info_dict
				# 方法2:pandas
				# names = ['comment', 'ip', 'protocol', 'port', 'username', 'password', 'secret', 'device_type']
				# df = pandas.read_excel(self.device_file, usecols='B:I', names=names, keep_default_na=False )
				# data = df.to_dict(orient='records')
				# for row in data:
				# 	row['cmd_list'] = self.get_cmd_info(row['device_type'])
				# 	yield  row
		except Exception as e:
			print('Error:', e)
		finally:
			wb.close()
	def get_cmd_info(self, cmd_sheet):
		"""获取命令信息，返回所有命令列表"""
		cmd_list = []
		try:
			for row in cmd_sheet.iter_rows(min_row=2, max_col=2):
				if str(row[0].value).strip() != '#' and row[1].value:
					# 跳过注释行，去掉命令左右空白
					cmd_list.append(row[1].value.strip())
			return cmd_list
		except Exception as e:
			print('get_cmd_info Error: ', e)
	def connectHandler(self, host):
		"""定义一个netmiko对象"""
		try:
			# 判断使用ssh协议
			if host['protocol'].lower().strip() == 'ssh':
				host['port'] = host['port'] if (host['port'] not in [22, None]) else 22
				host.pop('protocol'), host.pop('cmd_list')
				if 'huawei' in host['device_type']:
					# 华为超时设置为15秒
					connect = ConnectHandler(**host, conn_timeout=15)
				# elif 'fortinet' in host['device_type']:
					# 调用重写的MyFortinetSSH类
					# connect = MyFortinetSSH(**host)
				else:
					connect = ConnectHandler(**host)
			#判断使用telnet协议
			elif host['protocol'].lower().strip() == 'telnet':
				host['port'] = host['port'] if (host['port'] not in [23, None]) else 23
				host.pop('protocol'), host.pop('cmd_list')
				# netmiko里面支持telnet协议，设备类型格式加_telnet
				host['device_type'] = host['device_type'] + '_telnet'
				# fast_cli=False ，修复Telnet login authentication 报错。
				connect = ConnectHandler(**host, fast_cli=False)
			# 不支持的协议
			else:
				res = '{} 暂不支持{}协议！'.format(host['ip'], host['protocol'])
				raise ValueError(res)
			return connect
		# 捕获异常
		except NetMikoTimeoutException as e:
			res = 'Failed......{:<15} 连通性问题！'.format(host['ip'])
			print(res)
			self.write_to_file(**{'state': 0, 'error': str(res)})
		except AuthenticationException as e:
			res = 'Failed......{} 用户名或密码错误！'.format(host['ip'])
			print(res)
			self.write_to_file(**{'state': 0, 'error': str(res)})
		except SSHException as e:
			res = 'Failed......{} SSH版本不兼容！'.format(host['ip'])
			print(res)
			self.write_to_file(**{'state': 0, 'error': str(res)})
		except Exception as e:
			print('connectHandler Failed: {}'.format(e))
			self.write_to_file(**{'state': 0, 'error': str(res)})
	def run_cmd(self, host, cmds, enable=False):
		"""执行命令，保存信息"""
		# 特权功能标识位
		enable = True if host['secret'] else False
		conn = self.connectHandler(host)
		if conn:
			# 获取到设备名称，不同人或不同厂商命名都会有些特殊，按需优化
			hostname = conn.find_prompt().strip('<').strip('>')
			dirname = host['ip'] + '_' + hostname
			dirpath = os.path.join(self.log, self.logtime, dirname)
			# 逐级创建目录
			try:
				os.makedirs(dirpath)
			except:
				raise Exception('文件夹创建失败！')
			try:
				if cmds:
					# 判断命令为真的条件
					for cmd in cmds:
						if enable:
							# 进入特权模式
							conn.enable()
							# output += conn.send_command(cmd, strip_command=False, strip_prompt=False)
							output = conn.send_command(cmd, strip_command=False, strip_prompt=False)
							# print(output)
							data = {'state': 1, 'result': output, 'path': os.path.join(dirpath, cmd + '.conf')}
							self.write_to_file(**data)
						else:
							# output += conn.send_command(cmd, strip_command=False, strip_prompt=False)
							output = conn.send_command(cmd, strip_command=False, strip_prompt=False)
							# print(output)
							data = {'state': 1, 'result': output, 'path': os.path.join(dirpath, cmd + '.conf')}
							self.write_to_file(**data)
				else:
					# 拓展用于ftp/sftp/scp备份使用
					pass
			except Exception as e:
				print(f'run_cmd Failed: {e}')
			finally:
				# 最后关闭会话
				conn.disconnect()
	def run_t(self, host):
		"""仅执行命令"""
		try:
			conn = self.connectHandler(host)
			if conn:
				# 获取到设备名称
				hostname = conn.find_prompt()
				output = '获取的设备提示符：{}'.format((hostname))
				# print(output)
				self.write_to_file(**{'state': 2, 'result': output})
				# 关闭会话
				conn.disconnect()
		except Exception as e:
			print(f'run_cmd Failed:{e}')
	def connect_t(self):
		"""连接测试"""
		start_time = datetime.now()
		hosts = self.get_device_info()
		for host in hosts:
			self.pool.apply_async(self.run_t, args=(host,))
		self.pool.close()
		self.pool.join()
		end_time = datetime.now()
		print('>>>>所有连接测试已经执行完成，总共耗时{:0.2f}秒.<<<'.format((end_time - start_time).total_seconds()))
	def connect(self):
		"""主程序"""
		start_time = datetime.now()
		#hosts 是一个生成器，需要for循环进行遍历
		hosts = self.get_device_info()
		for host in hosts:
			# self.run_cmd(host, host['cmd_list'])
			self.pool.apply_async(self.run_cmd, args=(host, host['cmd_list']))
		self.pool.close()
		self.pool.join()
		end_time = datetime.now()
		print('-' * 50)
		print('>>>>所有已经执行完成，总共耗时{:0.2f}秒.<<<'.format((end_time - start_time).total_seconds()))
if __name__=='__main__':
	# debug定位问题
	# logging.basicConfig(filename='debug.log', level=logging.DEBUG)
	# logging.getLogger('netmiko')
	# 执行主程序
	XunJian().connect()
	# 连接测试
	# XunJian().connect_t()
	# input("Press <enter> to quit")
